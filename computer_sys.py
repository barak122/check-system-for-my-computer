from openpyxl import *
from tkinter import *
from tkinter import messagebox
import cv2
import face_recognition
import numpy as np
from PIL import Image, ImageTk
import time

workbook = load_workbook(filename=r"C:\Users\97252\Desktop\אופיר\QTdesineir\מערכת כניסה\computer_sys.xlsx")
sheet = workbook.active
row = sheet.max_row

global user_name11
global another_win

win_vari = False


def face():
    global windd
    global win_vari
    video_capture = cv2.VideoCapture(0)
    ofir_image = face_recognition.load_image_file("ofir.jpg")
    ofir_face_encoding = face_recognition.face_encodings(ofir_image)[0]
    known_face_encoding = [ofir_face_encoding]
    known_face_names = ["Ofir"]
    while True:
        ret, frame = video_capture.read()
        rgb_frame = frame[:, :, ::-1]
        face_locations = face_recognition.face_locations(rgb_frame)
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)
        for (top, right, bottum, left), face_encoding in zip(face_locations, face_encodings):  # TODO לכל פרצוף בווידאו
            matches = face_recognition.compare_faces(known_face_encoding,
                                                     face_encoding)  # TODO זה עושה השוואה בין הפרצופים שידועים לנו לבין הפרצופים הווידאו
            name = "UnKnown"
            faces_distances = face_recognition.face_distance(known_face_encoding,
                                                             face_encoding)  # TODO כמה זה קרוב או רחוק לפרצוף קיים
            best_match_index = np.argmin(faces_distances)
            if matches[best_match_index]:  # TODO אם הפרצוף שנמצא קיים השם משונה לשם של אותו פרצוף
                name = known_face_names[best_match_index]
            cv2.rectangle(frame, (left, top), (right, bottum), (34, 177, 76),
                          (3))  # TODO זה עושה ריבוע סביב הפרצוף(אחרי הגדרה של איפה הפרצוף והגדרה של הצבע והעובי)
            cv2.rectangle(frame, (left, bottum - 35), (right, bottum), (34, 177, 76),
                          cv2.FILLED)  # TODO זה עושה ריבוע סביב השם
            cv2.putText(frame, name, (left + 6, bottum - 6), cv2.FONT_HERSHEY_DUPLEX, 1.0, (
                255, 255, 255))  # TODO זה איפה יהיה הריבוע שאמור להיות סביב השם , הצבע שלו העובי והצבע של הטקסט בתוכו
            if name == "Ofir":
                windd()
                win_vari = True
            else:
                break
        # cv2.imshow('Video', frame)  # TODO זה מראה את הווידאו
        if cv2.waitKey(2) & 0xFF == ord("q"):  # TODO  זה יוצא מהלולאה "q" אם אני לוחץ על
            break


win = Tk()
win.title("אישור כניסה")
win.attributes('-fullscreen', True)

win['background'] = '#427ec6'
path = r"C:\Users\97252\Desktop\אופיר\QTdesineir\מערכת כניסה\face_scan.png"
img = ImageTk.PhotoImage(Image.open(path))
panel = Label(win, image=img)
panel.place(x=450, y=20)


def windd():
    wind = Tk()
    wind.attributes('-fullscreen', True)
    wind['background'] = '#427ec6'
    wind.title("חלון כניסה")

    label1 = Label(wind, text="כאן תוכל להכנס למערכת", font=("Arial", 40), bg='#427ec6', fg="#000000", width="32",
                   height="2").place(x=270, y=30)
    label2 = Label(wind, text=":הכנס כאן שם משתמש וסיסמא", font=("Arial", 40), bg='#427ec6', fg="#000000", width="32",
                   height="3").place(x=560, y=110)
    label3 = Label(wind, text=":שם משתמש", font=("Arial", 32), bg='#427ec6', fg="#000000", width="32",
                   height="2").place(x=975, y=260)
    label4 = Label(wind, text=":סיסמא", font=("Arial", 32), bg='#427ec6', fg="#000000", width="32",
                   height="2").place(x=975, y=460)

    var1 = StringVar()
    var2 = IntVar()

    def error():
        messagebox.showerror("חלון שגיאה", "השם משתמש או הסיסמא לא נכונים")
        another_win()

    def another_win():
        another_win = Tk()
        another_win.title("חלון רישום")
        another_win.attributes('-fullscreen', True)
        another_win['background'] = '#427ec6'

        label1 = Label(another_win, text="כאן תוכלו להירשם", font=("Arial", 40), bg='#427ec6', fg="#000000", width="32",
                       height="2").place(x=270, y=30)
        label2 = Label(another_win, text=":הירשמו עם שם משתמש וסיסמא", font=("Arial", 40), bg='#427ec6', fg="#000000",
                       width="32",
                       height="3").place(x=560, y=110)
        label3 = Label(another_win, text=":שם משתמש", font=("Arial", 32), bg='#427ec6', fg="#000000", width="32",
                       height="2").place(x=975, y=260)
        label4 = Label(another_win, text=":סיסמא", font=("Arial", 32), bg='#427ec6', fg="#000000", width="32",
                       height="2").place(x=975, y=460)

        var11 = StringVar()
        var22 = StringVar()

        user_name11 = Entry(another_win, textvariable=var11)
        user_name11.place(x=1015, y=302)
        password11 = Entry(another_win, textvariable=var22)
        password11.place(x=1015, y=502)

        def excel_in():
            global labelh1
            workbook = load_workbook(filename=r"C:\Users\97252\Desktop\אופיר\QTdesineir\מערכת כניסה\computer_sys.xlsx")
            # ws = workbook.worksheets[0]
            sheet = workbook.active
            names = []

            row = sheet.max_row
            oenter = False
            for i in range(1, row + 1):
                names.append(sheet.cell(i, 1).value)
            for i in range(row):
                if len(user_name11.get()) >= 3:
                    if user_name11.get() not in names:
                        if len(password11.get()) >= 8:
                            sheet[f"A{row + 1}"] = user_name11.get()
                            sheet[f"B{row + 1}"] = password11.get()
                            workbook.save(r"C:\Users\97252\Desktop\אופיר\QTdesineir\מערכת כניסה\computer_sys.xlsx")
                            if oenter:
                                labelh1.config(text="!פרטיך נשמרו בהצלחה")
                            else:
                                labelh11 = Label(another_win, text="!פרטיך נשמרו בהצלחה", font=("Arial", 12),
                                                 bg="#00fc00", fg="#000000", width="30",
                                                 height="2").place(x=440, y=490)
                        else:
                            labelh1 = Label(another_win, text="!הסיסמא חייבת להכיל לפחות 8 תווים", font=("Arial", 12),
                                            bg='#fff200', fg="#000000", width="30",
                                            height="2").place(x=440, y=490)
                    else:
                        labelh1 = Label(another_win, text="!השם משתמש קיים כבר", font=("Arial", 12),
                                        bg='#fff200', fg="#000000", width="30",
                                        height="2").place(x=440, y=490)
                        oenter = True
                else:
                    label = Label(another_win, text="!השם משתמש חייב להכיל לכל הפחות 3 תווים", font=("Arial", 12),
                                  bg='#ff0000', fg="#000000", width="30",
                                  height="2").place(x=440, y=490)

        button11 = Button(another_win, text="לחץ כאן על מנת לסיים את ההרשמה", font=("Arial", 16), bg='#7bbbbf',
                          fg="#000000",
                          width="30",
                          height="2", command=excel_in).place(x=180, y=690)
        button223 = Button(another_win, text="לחלון התחברות", font=("Arial", 16), bg='#7bbbbf', fg="#000000",
                           width="30",
                           height="2", command=windd).place(x=580, y=690)

        another_win.mainloop()

    def excel_check():
        nampass = {}
        for i in range(1, row):
            nampass[sheet.cell(i, 1).value] = sheet.cell(i, 2).value
        print(nampass)
        for i in range(row):
            for key, value in nampass.items():
                if key == user_name.get() and value == password.get():
                    print("succeed");
                    wind.destroy()
                else:
                    error()

    user_name = Entry(wind, textvariable=var1)
    user_name.place(x=1015, y=302)
    password = Entry(wind, textvariable=var2)
    password.place(x=1015, y=502)

    button = Button(wind, text="לחץ כאן על מנת להתחבר", font=("Arial", 16), bg='#7bbbbf', fg="#000000", width="20",
                    height="2", command=excel_check).place(x=730, y=690)

    button22 = Button(wind, text="לחץ כאן על מנת להירשם", font=("Arial", 16), bg='#7bbbbf',
                      fg="#000000", width="20",
                      height="2", command=another_win).place(x=340, y=690)

    wind.mainloop()


button_face = Button(win, text="!לחץ כאן בבקשה ולאחר מכן חכה רגע ואל תזוז", font=("Arial", 40), bg="#34575a",
                     fg="#000000", width="32",
                     height="2", command=face).place(x=270, y=600)  # TODO לשנות שלפרצוף שלי הוא ייצא
if win_vari:
    print("00035")
    win.destroy()

win.mainloop()
video_capture.release()

first_win()
