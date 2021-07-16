import wmi
from openpyxl import *
from openpyxl.styles import Font, PatternFill
import time

# Initializing the wmi constructor
f = wmi.WMI()

things_open = []
wb1 = load_workbook(r"C:\Users\97252\Desktop\thing_not_in_hebrew\computer_sys20.xlsx")
ws1 = wb1.active
row = ws1.max_row
sheets = wb1.sheetnames
worksheet.views.rightToLeft = True


# אם יש משתמש שלא קיים לו גיליון הפונקצייה הזאת יורת לו גיליון חדש עם השם שלו ועוד כמה דברים במידה והיא רואה שבאמת לא קיים לו גיליון
def create_sheet_of_user():
    global row, wb1, ws1, sheets
    users = []
    for e in range(2, row + 1):
        users.append(ws1.cell(e, 1).value)
    print(f"Users {users}")
    for r in users:
        if r not in sheets:
            print("r not in sheets")
            wb1.create_sheet(r, len(sheets))
            ws1_new = wb1[r]
            ws1_new["A1"] = "אפליקצייה שהייתה בשימוש"
            print("עמודת אפליקצייה שהייתה בשימוש")
            ws1_new.row_dimensions[1].height = 20
            # ws1_new.column_dimensions[1].width = 45
            print("Longer cell")
            ws1_new["B1"] = "תאריך"
            print("עמודת תאריך")
            ws1_new["C1"] = "שעה"
            print("עמודת שעה")
            ws1_new['A1'].font = Font(size=16, bold=True)
            print("font - 1")
            ws1_new['B1'].font = Font(size=16, bold=True)
            print("font - 2")
            ws1_new['C1'].font = Font(size=16, bold=True)
            print("font - 3")
            # wb1.save("computer_sys1.xlsx")
            wb1.save(r"C:\Users\97252\Desktop\thing_not_in_hebrew\computer_sys20.xlsx")
            print("Success")


create_sheet_of_user()

# מכניס לרשימה את שמות כל האפליקציות שפתוחות כעת במחשב
for process in f.Win32_Process():
    things_open.append(process.Name)
redFill = PatternFill(start_color='00FF0000',
                      end_color='00FF0000',
                      fill_type='solid')
while True:
    things_open2 = []
    # מכניס לרשימה את שמות כל האפליקציות שפתוחות כעת במחשב , וזה בודק האם יש שינוי בין הרשימה שנאספה מראש לרשימה
    # "החדשה" אם כן הוא מוצא מה לא היה קודם
    for process in f.Win32_Process():
        things_open2.append(process.Name)
    for i in things_open2:
        if i not in things_open:
            print(i)
            things_open = things_open2
            user_cell = ""
            for c in range(2, row + 1):
                if ws1[f'C{c}'].fill == redFill:  # בודק האם התא ממולא בצבע אדום או שלא
                    pass
                else:
                    user_cell = ws1.cell(c, 1).value
                    print(f"The cell is not red {user_cell}")
                    break
            new_sheet = wb1[user_cell]
            print(new_sheet)
            print(new_sheet.max_row)
            x = new_sheet.max_row
            new_sheet.cell(x+1, 1).value = i
            print("New cell app was used")
            new_sheet.cell(x+1, 2).value = time.strftime("%d/%m/%y")
            print("New cell date")
            new_sheet.cell(x+1, 3).value = time.strftime("%H:%M:%S")
            print("New cell time")
            # wb1.save("computer_sys20.xlsx")
            wb1.save(r"C:\Users\97252\Desktop\thing_not_in_hebrew\computer_sys20.xlsx")
            print("Very good , save")
